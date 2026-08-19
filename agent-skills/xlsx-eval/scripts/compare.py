#!/usr/bin/env python3
# compare.py — xlsx-eval skill (v2)
#   Compare a submitted workbook against an expected (answer-key) workbook and
#   report formula / value / structural differences with a match score.
#
#   python compare.py --expected key.xlsx --submitted cand.xlsx
#       [--json] [--max 100] [--tol 1e-9] [--ignore-case] [--grade]
#
# v2 additions over the original:
#   --tol T          numeric tolerance so 3.14159 vs 3.14 (or float noise) can match
#   --ignore-case    normalize formula case + whitespace before comparing
#   --grade          concise scored output (per-sheet + overall), for grading tasks
#   named ranges     defined-name (named range) diffs are reported
#   value typing     numeric compare is tolerance-aware; text compare is exact
import argparse, json, re, sys


def _norm_formula(f, ignore_case):
    if f is None:
        return None
    s = f
    if ignore_case:
        # collapse whitespace and uppercase — Excel formulas are case-insensitive
        s = re.sub(r"\s+", "", s).upper()
    return s


def load(path):
    from openpyxl import load_workbook
    wbf = load_workbook(path, data_only=False)
    try:
        wbv = load_workbook(path, data_only=True)
    except Exception:
        wbv = None
    out = {}
    for name in wbf.sheetnames:
        wsf = wbf[name]
        wsv = wbv[name] if (wbv and name in wbv.sheetnames) else None
        cells = {}
        for row in wsf.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                formula = c.value if (isinstance(c.value, str) and c.value.startswith("=")) else None
                value = wsv[c.coordinate].value if (formula is not None and wsv is not None) else c.value
                cells[c.coordinate] = {"formula": formula, "value": value, "fmt": c.number_format}
        out[name] = cells
    # defined names (named ranges) — workbook-level
    named = {}
    try:
        for nm in wbf.defined_names:
            dn = wbf.defined_names[nm]
            named[nm] = dn.value
    except Exception:
        # openpyxl <3.1 exposed defined_names.definedName; be forgiving
        try:
            for dn in wbf.defined_names.definedName:
                named[dn.name] = dn.value
        except Exception:
            pass
    return {"sheets": out, "named": named}


def _values_match(ev, sv, tol):
    """Tolerance-aware value compare. Numbers within tol match; else exact str."""
    if ev is None and sv is None:
        return True
    # try numeric
    try:
        ef, sf = float(ev), float(sv)
        if abs(ef - sf) <= tol or (ef != 0 and abs((ef - sf) / ef) <= tol):
            return True
        return False
    except (TypeError, ValueError):
        pass
    return str(ev) == str(sv)


def compare(expected, submitted, mx, tol, ignore_case, check_format=False, weights=None):
    exp_sheets, sub_sheets = expected["sheets"], submitted["sheets"]
    sheets = []
    total = matched = 0
    # Weighted scoring: weights maps sheet name -> points (or "*" -> default per sheet).
    # Points earned = sheetPct/100 * sheet_weight; possible = sum of weights. When no
    # weights given, we fall back to raw cell counts (every cell equal).
    weighted = bool(weights)
    default_w = (weights or {}).get("*")
    pts_earned = pts_possible = 0.0
    for name in exp_sheets:
        ecells = exp_sheets[name]
        scells = sub_sheets.get(name, {})
        fdiffs, vdiffs, missing, extra, fmtdiffs = [], [], [], [], []
        for ref, ev in ecells.items():
            total += 1
            sv = scells.get(ref)
            if sv is None:
                missing.append(ref)
                continue
            ef = _norm_formula(ev.get("formula"), ignore_case)
            sf = _norm_formula(sv.get("formula"), ignore_case)
            if ef != sf:
                fdiffs.append({"ref": ref, "expected": ev.get("formula"), "submitted": sv.get("formula")})
                continue
            if not _values_match(ev.get("value"), sv.get("value"), tol):
                vdiffs.append({"ref": ref, "expected": ev.get("value"), "submitted": sv.get("value")})
                continue
            # value+formula matched; optionally check number format too
            if check_format and ev.get("fmt") != sv.get("fmt"):
                fmtdiffs.append({"ref": ref, "expected": ev.get("fmt"), "submitted": sv.get("fmt")})
                continue
            matched += 1
        for ref in scells:
            if ref not in ecells:
                extra.append(ref)
        cellsInSheet = len(ecells)
        bad = len(fdiffs) + len(vdiffs) + len(missing) + (len(fmtdiffs) if check_format else 0)
        okInSheet = cellsInSheet - bad
        sheetPct = round(100.0 * okInSheet / cellsInSheet, 1) if cellsInSheet else 100.0
        # weighted contribution for this sheet
        if weighted:
            w = weights.get(name, default_w)
            if w is not None:
                pts_possible += w
                pts_earned += (okInSheet / cellsInSheet if cellsInSheet else 1.0) * w
        row = {
            "sheet": name,
            "present": name in sub_sheets,
            "cells": cellsInSheet,
            "sheetPct": sheetPct,
            "formulaDiffs": fdiffs[:mx], "formulaDiffCount": len(fdiffs),
            "valueDiffs": vdiffs[:mx], "valueDiffCount": len(vdiffs),
            "missing": missing[:mx], "missingCount": len(missing),
            "extra": extra[:mx], "extraCount": len(extra),
        }
        if check_format:
            row["formatDiffs"] = fmtdiffs[:mx]
            row["formatDiffCount"] = len(fmtdiffs)
        if weighted:
            row["weight"] = weights.get(name, default_w)
        sheets.append(row)
    extra_sheets = [n for n in sub_sheets if n not in exp_sheets]

    # named ranges (defined names)
    en, sn = expected["named"], submitted["named"]
    name_diffs = []
    for k, ev in en.items():
        if k not in sn:
            name_diffs.append({"name": k, "expected": ev, "submitted": None})
        elif str(sn[k]) != str(ev):
            name_diffs.append({"name": k, "expected": ev, "submitted": sn[k]})
    named_missing = [k for k in en if k not in sn]

    pct = round(100.0 * matched / total, 1) if total else 100.0
    result = {
        "matchPct": pct, "cellsChecked": total, "cellsMatched": matched,
        "extraSheets": extra_sheets, "sheets": sheets,
        "namedRangeDiffs": name_diffs, "namedRangeMissing": named_missing,
    }
    if weighted:
        result["weightedScore"] = round(pts_earned, 2)
        result["weightedPossible"] = round(pts_possible, 2)
        result["weightedPct"] = round(100.0 * pts_earned / pts_possible, 1) if pts_possible else 100.0
    return result


def render(r, grade):
    lines = [f"Match: {r['matchPct']}%  ({r['cellsMatched']}/{r['cellsChecked']} cells)"]
    if "weightedPct" in r:
        lines.append(f"Weighted score: {r['weightedPct']}%  ({r['weightedScore']}/{r['weightedPossible']} pts)")
    if r["extraSheets"]:
        lines.append("Extra sheets in submission: " + ", ".join(r["extraSheets"]))
    if r["namedRangeDiffs"]:
        lines.append(f"Named-range diffs: {len(r['namedRangeDiffs'])}")
        if not grade:
            for d in r["namedRangeDiffs"]:
                lines.append(f"  {d['name']}: expected {d['expected']!r} got {d['submitted']!r}")
    for s in r["sheets"]:
        if not s["present"]:
            lines.append(f"\n[{s['sheet']}] MISSING from submission")
            continue
        fmtc = s.get("formatDiffCount", 0)
        wtag = f" · w={s['weight']}" if ("weight" in s and s["weight"] is not None) else ""
        if grade:
            flag = "✓" if s["sheetPct"] == 100.0 else " "
            fmtpart = f" {fmtc}fmt" if "formatDiffCount" in s else ""
            lines.append(f"  {flag} [{s['sheet']}] {s['sheetPct']}%  "
                         f"({s['formulaDiffCount']}f {s['valueDiffCount']}v {s['missingCount']}m {s['extraCount']}x{fmtpart}){wtag}")
            continue
        if not (s["formulaDiffCount"] or s["valueDiffCount"] or s["missingCount"] or s["extraCount"] or fmtc):
            lines.append(f"\n[{s['sheet']}] ✓ exact match{wtag}")
            continue
        parts = f"{s['formulaDiffCount']} formula, {s['valueDiffCount']} value, {s['missingCount']} missing, {s['extraCount']} extra"
        if "formatDiffCount" in s:
            parts += f", {fmtc} format"
        lines.append(f"\n[{s['sheet']}] {parts}{wtag}")
        for d in s["formulaDiffs"]:
            lines.append(f"  {d['ref']} formula: expected {d['expected']!r}  got {d['submitted']!r}")
        for d in s["valueDiffs"]:
            lines.append(f"  {d['ref']} value: expected {d['expected']!r}  got {d['submitted']!r}")
        for d in s.get("formatDiffs", []):
            lines.append(f"  {d['ref']} format: expected {d['expected']!r}  got {d['submitted']!r}")
        if s["missing"]:
            lines.append("  missing: " + ", ".join(s["missing"]))
        if s["extra"]:
            lines.append("  extra: " + ", ".join(s["extra"]))
    return "\n".join(lines)


def main():
    ap = argparse.ArgumentParser(description="Compare a submitted .xlsx against an expected/answer-key .xlsx.")
    ap.add_argument("--expected", required=True)
    ap.add_argument("--submitted", required=True)
    ap.add_argument("--json", action="store_true", help="machine-readable output")
    ap.add_argument("--max", type=int, default=100, help="cap reported diffs per sheet")
    ap.add_argument("--tol", type=float, default=1e-9, help="numeric tolerance (abs or rel) for value matches")
    ap.add_argument("--ignore-case", action="store_true", help="normalize formula case + whitespace")
    ap.add_argument("--grade", action="store_true", help="concise per-sheet scored output")
    ap.add_argument("--check-format", action="store_true", help="also compare each cell's number format")
    ap.add_argument("--weights", help="JSON file mapping sheet name -> points (use \"*\" for a per-sheet default); score becomes weighted")
    a = ap.parse_args()
    weights = None
    if a.weights:
        try:
            with open(a.weights) as fh:
                weights = json.load(fh)
            if not isinstance(weights, dict):
                raise ValueError("weights JSON must be an object of sheet -> points")
        except Exception as e:
            msg = {"ok": False, "error": "weights: " + str(e)}
            print(json.dumps(msg) if a.json else "error: " + str(e)); sys.exit(1)
    try:
        r = compare(load(a.expected), load(a.submitted), a.max, a.tol, a.ignore_case,
                    check_format=a.check_format, weights=weights)
    except Exception as e:
        msg = {"ok": False, "error": str(e)}
        print(json.dumps(msg) if a.json else "error: " + str(e)); sys.exit(1)
    if a.json:
        print(json.dumps({"ok": True, **r}))
    else:
        print(render(r, a.grade))


if __name__ == "__main__":
    main()
