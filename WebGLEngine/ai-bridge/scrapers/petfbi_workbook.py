#!/usr/bin/env python3
# petfbi_workbook.py - create / manage the PetFBI .xlsx workbook used by the
# posting pipeline. Drives the setup wizard (state + FB page name + page id ->
# "<STATE> Pages" sheet), admin management (Admins sheet), and the template
# editor (Templates sheet). Each subcommand prints one JSON line to stdout.
#
#   python petfbi_workbook.py status      --wb PATH
#   python petfbi_workbook.py init        --wb PATH [--state TX]
#   python petfbi_workbook.py add-page    --wb PATH --state TX --fbPage "Name" [--pageId N --url U --businessId B --town T --kind statewide]
#   python petfbi_workbook.py add-admin   --wb PATH --name "X" --email a@b.c [--towns "a,b" --state TX]
#   python petfbi_workbook.py list-pages  --wb PATH [--state TX]
#   python petfbi_workbook.py list-admins --wb PATH
#   python petfbi_workbook.py get-templates --wb PATH
#   python petfbi_workbook.py set-templates --wb PATH --data '{"Lost":{"Dog":"..."}}'
#   python petfbi_workbook.py del-row     --wb PATH --sheet "TX Pages" --row 3
import sys, os, json, argparse

PAGE_HEADERS = ["fbPage", "url", "pageId", "businessId", "town", "kind"]
ADMIN_HEADERS = ["Name", "Email", "Towns", "State"]
_LOST = ">>  [TOWN], [STATE] - Lost [COLOR] {kind}\n\n[STREET]\n\nLost on [DATE]. [GENDER-HE/SHE] is [BREED], age [AGE]. Collar: [COLLAR].\n\n[DESCRIPTION]\n\nPlease share and contact [EMAIL] if you can help bring [GENDER-HER/HIS] home.\n\nPetFBI: https://petfbi.org/api/view/[PETIDNUM]"
_FOUND = "<<  [TOWN], [STATE] - Found [COLOR] {kind}\n\n[STREET]\n\nFound on [DATE]. Collar: [COLLAR].\n\n[DESCRIPTION]\n\nPlease share and contact [EMAIL] if this is your {low}.\n\nPetFBI: https://petfbi.org/api/view/[PETIDNUM]"
DEFAULT_TEMPLATES = {
    "Lost": {"Dog": _LOST.format(kind="Dog"), "Cat": _LOST.format(kind="Cat")},
    "Found": {"Dog": _FOUND.format(kind="Dog", low="dog"), "Cat": _FOUND.format(kind="Cat", low="cat")},
}


def out(d):
    sys.stdout.write(json.dumps(d))
    sys.exit(0)


def _op():
    try:
        import openpyxl
        return openpyxl
    except Exception as e:
        out({"ok": False, "error": "openpyxl not installed - run: pip install openpyxl  (" + str(e) + ")"})


def state_sheet(state):
    return (state or "").strip().upper() + " Pages"


def load_wb(path, create=False):
    op = _op()
    if os.path.exists(path):
        return op.load_workbook(path)
    if create:
        wb = op.Workbook()
        wb.remove(wb.active)
        return wb
    return None


def ensure_sheet(wb, name, headers):
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(name)
    ws.append(headers)
    return ws


def _rows(ws):
    return list(ws.iter_rows(values_only=True))


def cmd_status(a):
    op = _op()
    if not os.path.exists(a.wb):
        out({"ok": True, "exists": False, "path": a.wb, "sheets": [], "pages": 0, "admins": 0, "states": []})
    wb = op.load_workbook(a.wb, read_only=True)
    sheets = wb.sheetnames
    pages, states = 0, []
    for s in sheets:
        if s.endswith(" Pages"):
            states.append(s[:-6])
            pages += max(0, wb[s].max_row - 1)
    admins = max(0, wb["Admins"].max_row - 1) if "Admins" in sheets else 0
    out({"ok": True, "exists": True, "path": a.wb, "sheets": sheets, "pages": pages, "admins": admins, "states": states})


def cmd_init(a):
    wb = load_wb(a.wb, create=True)
    ensure_sheet(wb, "Admins", ADMIN_HEADERS)
    tws = ensure_sheet(wb, "Templates", ["Status", "Type", "Template"])
    if tws.max_row <= 1:
        for st, types in DEFAULT_TEMPLATES.items():
            for ty, tpl in types.items():
                tws.append([st, ty, tpl])
    if a.state:
        ensure_sheet(wb, state_sheet(a.state), PAGE_HEADERS)
    wb.save(a.wb)
    cmd_status(a)


def cmd_add_page(a):
    wb = load_wb(a.wb, create=True)
    ws = ensure_sheet(wb, state_sheet(a.state), PAGE_HEADERS)
    ws.append([a.fbPage, a.url or "", a.pageId or "", a.businessId or "", a.town or "", a.kind or "statewide"])
    wb.save(a.wb)
    out({"ok": True, "added": {"state": a.state, "fbPage": a.fbPage, "pageId": a.pageId or ""}})


def cmd_add_admin(a):
    wb = load_wb(a.wb, create=True)
    ws = ensure_sheet(wb, "Admins", ADMIN_HEADERS)
    ws.append([a.name, a.email, a.towns or "", a.state or ""])
    wb.save(a.wb)
    out({"ok": True, "added": {"name": a.name, "email": a.email}})


def cmd_list_pages(a):
    op = _op()
    if not os.path.exists(a.wb):
        out({"ok": True, "pages": []})
    wb = op.load_workbook(a.wb, read_only=True)
    pages = []
    for s in wb.sheetnames:
        if not s.endswith(" Pages"):
            continue
        st = s[:-6]
        if a.state and st.upper() != a.state.strip().upper():
            continue
        rows = _rows(wb[s])
        if not rows:
            continue
        hdr = [str(c or "") for c in rows[0]]
        for ri, r in enumerate(rows[1:], start=2):
            d = {hdr[i]: (r[i] if i < len(r) else "") for i in range(len(hdr))}
            d["state"] = st
            d["_row"] = ri
            d["_sheet"] = s
            pages.append(d)
    out({"ok": True, "pages": pages})


def cmd_list_admins(a):
    op = _op()
    if not os.path.exists(a.wb):
        out({"ok": True, "admins": []})
    wb = op.load_workbook(a.wb, read_only=True)
    if "Admins" not in wb.sheetnames:
        out({"ok": True, "admins": []})
    rows = _rows(wb["Admins"])
    hdr = [str(c or "") for c in rows[0]] if rows else ADMIN_HEADERS
    admins = []
    for ri, r in enumerate(rows[1:], start=2):
        d = {hdr[i]: (r[i] if i < len(r) else "") for i in range(len(hdr))}
        d["_row"] = ri
        admins.append(d)
    out({"ok": True, "admins": admins})


def cmd_get_templates(a):
    op = _op()
    tpl = {}
    if os.path.exists(a.wb):
        wb = op.load_workbook(a.wb, read_only=True)
        if "Templates" in wb.sheetnames:
            for r in _rows(wb["Templates"])[1:]:
                if r and r[0] and r[1]:
                    tpl.setdefault(str(r[0]), {})[str(r[1])] = str(r[2] or "")
    if not tpl:
        tpl = DEFAULT_TEMPLATES
    out({"ok": True, "templates": tpl})


def cmd_set_templates(a):
    data = json.loads(a.data) if a.data else json.load(open(a.json))
    wb = load_wb(a.wb, create=True)
    if "Templates" in wb.sheetnames:
        del wb["Templates"]
    ws = wb.create_sheet("Templates")
    ws.append(["Status", "Type", "Template"])
    n = 0
    for st, types in data.items():
        for ty, t in types.items():
            ws.append([st, ty, t])
            n += 1
    wb.save(a.wb)
    out({"ok": True, "count": n})


def cmd_del_row(a):
    op = _op()
    if not os.path.exists(a.wb):
        out({"ok": False, "error": "no workbook"})
    wb = op.load_workbook(a.wb)
    if a.sheet not in wb.sheetnames:
        out({"ok": False, "error": "no sheet " + a.sheet})
    wb[a.sheet].delete_rows(int(a.row), 1)
    wb.save(a.wb)
    out({"ok": True, "deleted": {"sheet": a.sheet, "row": a.row}})


def main():
    p = argparse.ArgumentParser()
    sub = p.add_subparsers(dest="cmd")
    for name in ("status", "init", "add-page", "add-admin", "list-pages", "list-admins", "get-templates", "set-templates", "del-row"):
        sp = sub.add_parser(name)
        sp.add_argument("--wb", required=True)
        sp.add_argument("--state", default="")
        sp.add_argument("--fbPage", default="")
        sp.add_argument("--pageId", default="")
        sp.add_argument("--url", default="")
        sp.add_argument("--businessId", default="")
        sp.add_argument("--town", default="")
        sp.add_argument("--kind", default="")
        sp.add_argument("--name", default="")
        sp.add_argument("--email", default="")
        sp.add_argument("--towns", default="")
        sp.add_argument("--data", default="")
        sp.add_argument("--json", default="")
        sp.add_argument("--sheet", default="")
        sp.add_argument("--row", default="")
    a = p.parse_args()
    fn = {
        "status": cmd_status, "init": cmd_init, "add-page": cmd_add_page, "add-admin": cmd_add_admin,
        "list-pages": cmd_list_pages, "list-admins": cmd_list_admins, "get-templates": cmd_get_templates,
        "set-templates": cmd_set_templates, "del-row": cmd_del_row,
    }.get(a.cmd)
    if not fn:
        out({"ok": False, "error": "unknown command"})
    try:
        fn(a)
    except SystemExit:
        raise
    except Exception as e:
        out({"ok": False, "error": str(e)})


if __name__ == "__main__":
    main()
